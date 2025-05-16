import happybase
from openpyxl import load_workbook
from tqdm import tqdm

# 连接到HBase
connection = happybase.Connection('localhost', port=9090)  # HBase服务器地址
connection.open()

# 创建表（如果尚未创建）
table_name = 'comments'
families = {'cf': dict()}  # 定义列族
if table_name not in connection.tables():
    connection.create_table(table_name, families)

# 打开Excel文件并读取数据
wb = load_workbook('comments.xlsx')
ws = wb.active

# 获取表对象
table = connection.table(table_name)

# 获取总行数（跳过表头）
total_rows = ws.max_row - 1

# 使用批量写入
with table.batch() as b:
    # 遍历Excel数据并存入HBase，使用tqdm显示进度
    for row in tqdm(ws.iter_rows(min_row=2, values_only=True), total=total_rows, desc="Writing to HBase"):
        try:
            row_data = {
                'cf:comment_user': str(row[1]),
                'cf:rating': str(row[2]),
                'cf:comment_time': str(row[3]),
                'cf:comment': str(row[4])
            }
            b.put(str(row[0]), row_data)  # 使用序号作为行键
        except Exception as e:
            print(f"Error processing row {row[0]}: {e}")

print("数据已成功存入HBase")
connection.close()